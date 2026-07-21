"""Create or safely update Jira tickets from the Excel workbook."""

from __future__ import annotations

import os
import sys
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any, Callable, Dict, List, Mapping, Optional, Sequence, Tuple

import requests
from openpyxl import load_workbook

from config_loader import ConfigError, load_legacy_config
from excel_workbook import set_jira_key_hyperlink
from ticket_schema import Action, normalize_issue_type
from validators import (
    TicketValidationError,
    is_blank,
    is_clear,
    resolve_action,
    validate_project_key,
    validate_ticket_row,
)


REQUIRED_COLUMNS = ["Summary", "IssueType", "JiraKey", "ErrorMessage"]
_REQUEST_ID_HEADERS = ("X-AREQUESTID", "X-Request-Id", "ATL-TraceId")
_EXPECTED_ROW_ERRORS = (
    KeyError,
    RuntimeError,
    TicketValidationError,
    ValueError,
    requests.RequestException,
)


@dataclass(frozen=True)
class IssueSnapshot:
    updated: str
    fields: Mapping[str, Any]


def load_config() -> Dict[str, Any]:
    """Keep the existing dictionary API while loading config.yaml centrally."""
    return load_legacy_config()


def jira_headers(config: Mapping[str, Any]) -> Dict[str, str]:
    return {
        "Authorization": f"Bearer {config['JIRA_TOKEN']}",
        "Accept": "application/json",
        "Content-Type": "application/json",
    }


def _request_id(response: requests.Response) -> str:
    for header in _REQUEST_ID_HEADERS:
        value = response.headers.get(header)
        if value:
            return str(value)[:100]
    return ""


def _raise_for_expected_status(
    response: requests.Response, expected: Sequence[int], operation: str
) -> None:
    if response.status_code in expected:
        return
    request_id = _request_id(response)
    suffix = f", Request-ID {request_id}" if request_id else ""
    raise RuntimeError(
        f"{operation} fehlgeschlagen: HTTP {response.status_code}{suffix}."
    )


def test_jira_connection(config: Dict[str, Any]) -> None:
    url = f"{config['JIRA_URL']}/rest/api/2/myself"
    response = requests.get(url, headers=jira_headers(config), timeout=30)
    _raise_for_expected_status(response, (200,), "Jira-Verbindung")
    user = response.json()
    if not isinstance(user, Mapping):
        raise RuntimeError("Jira-Benutzerantwort hat ein unerwartetes Format.")
    username = user.get("name") or user.get("key")
    if username:
        config["CURRENT_USER_NAME"] = str(username)
    display_name = user.get("displayName") or username or "unbekannt"
    print(f"  Verbunden als: {display_name}")


def test_project_access(config: Mapping[str, Any]) -> None:
    project_key = validate_project_key(
        config["PROJECT_KEY"], config["ALLOWED_PROJECTS"]
    )
    url = f"{config['JIRA_URL']}/rest/api/2/project/{project_key}"
    response = requests.get(url, headers=jira_headers(config), timeout=30)
    _raise_for_expected_status(response, (200,), f"Projektzugriff {project_key}")
    project = response.json()
    if not isinstance(project, Mapping):
        raise RuntimeError("Jira-Projektantwort hat ein unerwartetes Format.")
    print(f"  Projekt: {project.get('name', project_key)}")


def parse_labels(value: Optional[Any]) -> List[str]:
    if is_blank(value) or is_clear(value):
        return []
    return [label.strip() for label in str(value).split(",") if label.strip()]


def parse_components(value: Optional[Any]) -> List[Dict[str, str]]:
    if is_blank(value) or is_clear(value):
        return []
    return [
        {"name": name.strip()}
        for name in str(value).split(",")
        if name.strip()
    ]


def format_due_date(value: Any) -> Optional[str]:
    if is_blank(value) or is_clear(value):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    for pattern in ("%Y-%m-%d", "%d.%m.%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            continue
    raise ValueError(
        "Due Date muss ein echtes Excel-Datum oder DD.MM.YY/DD.MM.YYYY sein."
    )


def _set_optional_text(fields: Dict[str, Any], jira_field: str, value: Any) -> None:
    if is_blank(value):
        return
    fields[jira_field] = None if is_clear(value) else str(value).strip()


def build_issue_payload(
    config: Mapping[str, Any],
    row: Mapping[str, Any],
    include_project: bool = True,
    include_issue_type: Optional[bool] = None,
) -> Dict[str, Any]:
    """Build only known system fields; custom fields are metadata-driven later."""
    fields: Dict[str, Any] = {}
    include_issue_type = (
        include_project if include_issue_type is None else include_issue_type
    )

    if include_project:
        project_key = validate_project_key(
            row.get("Project"),
            config["ALLOWED_PROJECTS"],
            str(config["PROJECT_KEY"]),
        )
        fields["project"] = {"key": project_key}

    if include_issue_type:
        fields["issuetype"] = {"name": normalize_issue_type(row.get("IssueType"))}

    summary = row.get("Summary")
    if not is_blank(summary):
        if is_clear(summary):
            raise ValueError("Summary ist ein Pflichtfeld und kann nicht geleert werden.")
        fields["summary"] = str(summary).strip()

    _set_optional_text(fields, "description", row.get("Description"))

    priority = row.get("Priority")
    if not is_blank(priority):
        fields["priority"] = (
            None if is_clear(priority) else {"name": str(priority).strip()}
        )

    assignee = row.get("Assignee")
    if not is_blank(assignee):
        assignee_name = str(assignee).strip()
        if is_clear(assignee) or assignee_name.casefold() == "unassigned":
            fields["assignee"] = None
        elif assignee_name.casefold() == "assign to me":
            current_user = str(config.get("CURRENT_USER_NAME") or "").strip()
            if not current_user:
                raise ValueError(
                    "Assign to me erfordert den zuvor ermittelten Jira-Benutzer."
                )
            fields["assignee"] = {"name": current_user}
        else:
            fields["assignee"] = {"name": assignee_name}

    labels = row.get("Labels")
    if not is_blank(labels):
        fields["labels"] = parse_labels(labels)

    components = row.get("Components")
    if not is_blank(components):
        fields["components"] = parse_components(components)

    due_date = row.get("DueDate")
    if not is_blank(due_date):
        fields["duedate"] = format_due_date(due_date)

    return {"fields": fields}


def create_jira_issue(
    config: Mapping[str, Any], payload: Mapping[str, Any]
) -> Dict[str, Any]:
    url = f"{config['JIRA_URL']}/rest/api/2/issue"
    response = requests.post(
        url, headers=jira_headers(config), json=payload, timeout=30
    )
    _raise_for_expected_status(response, (200, 201), "Jira-Erstellung")
    result = response.json()
    if not isinstance(result, Mapping):
        raise RuntimeError("Jira-Erstellung hat ein unerwartetes Antwortformat.")
    return dict(result)


def update_jira_issue(
    config: Mapping[str, Any], jira_key: str, payload: Mapping[str, Any]
) -> None:
    url = f"{config['JIRA_URL']}/rest/api/2/issue/{jira_key}"
    response = requests.put(
        url, headers=jira_headers(config), json=payload, timeout=30
    )
    _raise_for_expected_status(
        response, (200, 204), f"Jira-Update {jira_key}"
    )


def fetch_issue_snapshot(
    config: Mapping[str, Any], jira_key: str, jira_fields: Sequence[str]
) -> IssueSnapshot:
    requested_fields = tuple(dict.fromkeys(tuple(jira_fields) + ("updated",)))
    url = f"{config['JIRA_URL']}/rest/api/2/issue/{jira_key}"
    response = requests.get(
        url,
        headers=jira_headers(config),
        params={"fields": ",".join(requested_fields)},
        timeout=30,
    )
    _raise_for_expected_status(response, (200,), f"Jira-Ticket laden {jira_key}")
    data = response.json()
    if not isinstance(data, Mapping):
        raise RuntimeError(f"Jira-Ticket {jira_key} hat ein unerwartetes Format.")
    fields = data.get("fields")
    if not isinstance(fields, Mapping):
        raise RuntimeError(f"Jira-Ticket {jira_key} enthaelt keine gueltigen Felder.")
    updated = str(fields.get("updated") or "").strip()
    if not updated:
        raise RuntimeError(f"Jira-Ticket {jira_key} enthaelt keinen updated-Zeitstempel.")
    return IssueSnapshot(updated=updated, fields=fields)


def _comparable_field_value(field_name: str, value: Any) -> Any:
    if field_name in {"priority", "assignee"}:
        return value.get("name") if isinstance(value, Mapping) else None
    if field_name == "components":
        if not isinstance(value, list):
            return ()
        return tuple(
            sorted(
                str(item.get("name"))
                for item in value
                if isinstance(item, Mapping) and item.get("name")
            )
        )
    if field_name == "labels":
        return tuple(sorted(str(item) for item in (value or [])))
    return value


def changed_update_fields(
    current_fields: Mapping[str, Any], desired_fields: Mapping[str, Any]
) -> Tuple[str, ...]:
    return tuple(
        field_name
        for field_name, desired in desired_fields.items()
        if _comparable_field_value(field_name, current_fields.get(field_name))
        != _comparable_field_value(field_name, desired)
    )


def confirm_update(jira_key: str, changed_fields: Sequence[str]) -> bool:
    field_list = ", ".join(changed_fields)
    try:
        answer = input(
            f"Update {jira_key} fuer Felder [{field_list}] bestaetigen? [ja/NEIN]: "
        )
    except EOFError:
        return False
    return answer.strip().casefold() in {"ja", "j", "yes", "y"}


def process_update(
    config: Mapping[str, Any],
    jira_key: str,
    payload: Mapping[str, Any],
    dry_run: bool,
    confirmation: Callable[[str, Sequence[str]], bool] = confirm_update,
) -> str:
    desired_fields = payload.get("fields")
    if not isinstance(desired_fields, Mapping):
        raise ValueError("Update-Payload enthaelt keine gueltigen fields.")
    snapshot = fetch_issue_snapshot(config, jira_key, tuple(desired_fields))
    changes = changed_update_fields(snapshot.fields, desired_fields)
    if not changes:
        print(f"  [NO CHANGE] {jira_key}: keine Aenderungen")
        return "NO_CHANGES"

    print(f"  [VORSCHAU] {jira_key}: {', '.join(changes)}")
    if dry_run:
        print(f"  [DRY RUN] Update {jira_key} validiert")
        return "DRY_RUN"
    if not confirmation(jira_key, changes):
        raise RuntimeError(f"Update {jira_key} wurde nicht bestaetigt.")

    latest = fetch_issue_snapshot(config, jira_key, ())
    if latest.updated != snapshot.updated:
        raise RuntimeError(
            f"Update {jira_key} blockiert: Ticket wurde zwischenzeitlich geaendert."
        )
    filtered_payload = {
        "fields": {field: desired_fields[field] for field in changes}
    }
    update_jira_issue(config, jira_key, filtered_payload)
    return "UPDATED"


def read_headers(sheet: Any) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for column_index, cell in enumerate(sheet[1], start=1):
        if cell.value:
            headers[str(cell.value).strip()] = column_index

    missing = [column for column in REQUIRED_COLUMNS if column not in headers]
    if "Action" not in headers and "Status" not in headers:
        missing.append("Action oder Status")
    if missing:
        raise ValueError(f"Fehlende Excel-Spalten: {', '.join(missing)}")
    return headers


def get_row_values(
    sheet: Any, row_index: int, headers: Mapping[str, int]
) -> Dict[str, Any]:
    return {
        column_name: sheet.cell(row=row_index, column=column_index).value
        for column_name, column_index in headers.items()
    }


def set_cell(
    sheet: Any,
    row_index: int,
    headers: Mapping[str, int],
    column: str,
    value: Any,
) -> None:
    sheet.cell(row=row_index, column=headers[column]).value = value


def _set_cell_if_present(
    sheet: Any,
    row_index: int,
    headers: Mapping[str, int],
    column: str,
    value: Any,
) -> None:
    if column in headers:
        set_cell(sheet, row_index, headers, column, value)


def validate_row(
    row: Mapping[str, Any],
    allowed_projects: Sequence[str] = ("DAH",),
    default_project: str = "DAH",
) -> List[str]:
    create_row = dict(row)
    create_row["Action"] = Action.CREATE.value
    return validate_ticket_row(create_row, allowed_projects, default_project)


def validate_edit_row(
    row: Mapping[str, Any],
    allowed_projects: Sequence[str] = ("DAH",),
    default_project: str = "DAH",
) -> List[str]:
    update_row = dict(row)
    update_row["Action"] = Action.UPDATE.value
    return validate_ticket_row(update_row, allowed_projects, default_project)


def _row_is_empty(row: Mapping[str, Any]) -> bool:
    return all(is_blank(value) for value in row.values())


def _record_success(
    sheet: Any,
    row_index: int,
    headers: Mapping[str, int],
    jira_url: str,
    jira_key: str,
    result: str,
) -> None:
    _set_cell_if_present(sheet, row_index, headers, "Action", Action.IGNORE.value)
    _set_cell_if_present(sheet, row_index, headers, "Status", "ALT")
    _set_cell_if_present(sheet, row_index, headers, "Result", result)
    _set_cell_if_present(sheet, row_index, headers, "ErrorMessage", "")
    _set_cell_if_present(
        sheet,
        row_index,
        headers,
        "CreatedAt",
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    )
    jira_key_cell = sheet.cell(row=row_index, column=headers["JiraKey"])
    set_jira_key_hyperlink(jira_key_cell, jira_url, jira_key)


def _record_error(
    sheet: Any,
    row_index: int,
    headers: Mapping[str, int],
    message: str,
) -> None:
    _set_cell_if_present(sheet, row_index, headers, "Status", "FEHLER")
    _set_cell_if_present(sheet, row_index, headers, "Result", "ERROR")
    _set_cell_if_present(sheet, row_index, headers, "ErrorMessage", message)


def _worksheet_by_name(workbook: Any, name: str) -> Optional[Any]:
    for index in range(1, workbook.Worksheets.Count + 1):
        sheet = workbook.Worksheets(index)
        if str(sheet.Name) == name:
            return sheet
    return None


def _com_header_map(sheet: Any) -> Dict[str, int]:
    max_column = max(1, int(sheet.UsedRange.Columns.Count))
    headers: Dict[str, int] = {}
    for column_index in range(1, max_column + 1):
        value = sheet.Cells(1, column_index).Value
        if value is not None and str(value).strip():
            headers[str(value).strip()] = column_index
    return headers


def reapply_dropdowns(excel_file: str) -> None:
    """Restore COM data validation after openpyxl saves an XLSM."""
    if not str(excel_file).lower().endswith(".xlsm"):
        return
    try:
        import win32com.client as win32
    except ImportError:
        print("  [WARN] pywin32 fehlt - Dropdowns nicht wiederhergestellt")
        return

    path = os.path.abspath(excel_file)
    excel = win32.gencache.EnsureDispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    workbook = None
    try:
        workbook = excel.Workbooks.Open(path)
        tickets = _worksheet_by_name(workbook, "Tickets")
        if tickets is None:
            raise RuntimeError("Blatt 'Tickets' fehlt; Dropdowns nicht anwendbar.")
        metadata = _worksheet_by_name(workbook, "Metadaten")
        if metadata is not None:
            source_headers = _com_header_map(metadata)
            source_names = {
                "Action": "Actions",
                "Project": "Projects",
                "IssueType": "IssueTypes",
                "Priority": "Priorities",
                "Components": "Components",
                "Status": "LegacyStatus",
            }
        else:
            metadata = _worksheet_by_name(workbook, "Listen")
            if metadata is None:
                raise RuntimeError("Blatt 'Metadaten' oder 'Listen' fehlt.")
            source_headers = {
                "Components": 1,
                "IssueTypes": 2,
                "Priorities": 3,
                "LegacyStatus": 4,
            }
            source_names = {
                "IssueType": "IssueTypes",
                "Priority": "Priorities",
                "Components": "Components",
                "Status": "LegacyStatus",
            }

        ticket_headers = _com_header_map(tickets)
        for target_name, source_name in source_names.items():
            target_column = ticket_headers.get(target_name)
            source_column = source_headers.get(source_name)
            if target_column is None or source_column is None:
                continue
            source_letter = _column_letter(source_column)
            target_letter = _column_letter(target_column)
            last_row = int(
                metadata.Cells(metadata.Rows.Count, source_column).End(-4162).Row
            )
            last_row = max(last_row, 2)
            formula = (
                f"='{metadata.Name}'!${source_letter}$2:${source_letter}${last_row}"
            )
            cell_range = tickets.Range(
                f"{target_letter}2:{target_letter}1000"
            )
            cell_range.Validation.Delete()
            cell_range.Validation.Add(3, 1, 1, formula)
            cell_range.Validation.InCellDropdown = True
            cell_range.Validation.IgnoreBlank = True
        workbook.Save()
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=False)
        excel.Quit()
    print("  Dropdowns wiederhergestellt.")


def _column_letter(index: int) -> str:
    letters = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def main() -> None:
    config = load_config()
    dry_run = str(config["DRY_RUN"]).lower() == "true"

    print("=" * 50)
    print("Jira Ticket-Verarbeitung aus Excel")
    print("=" * 50)
    print(f"  Jira-URL:    {config['JIRA_URL']}")
    print(f"  Projekt:     {config['PROJECT_KEY']}")
    print(f"  Excel:       {config['EXCEL_FILE']}")
    print(f"  Dry Run:     {dry_run}")
    print()

    print("Teste Jira-Verbindung...")
    test_jira_connection(config)
    print("Teste Projektzugriff...")
    test_project_access(config)
    print()

    excel_file = str(config["EXCEL_FILE"])
    keep_vba = excel_file.lower().endswith(".xlsm")
    workbook = load_workbook(excel_file, keep_vba=keep_vba)
    sheet = workbook["Tickets"] if "Tickets" in workbook.sheetnames else workbook.active
    headers = read_headers(sheet)
    allowed_projects = tuple(config["ALLOWED_PROJECTS"])
    default_project = str(config["PROJECT_KEY"])
    legacy_default_project = default_project if "Project" not in headers else None

    processed = created = updated = failed = skipped = 0
    try:
        for row_index in range(2, sheet.max_row + 1):
            row = get_row_values(sheet, row_index, headers)
            if _row_is_empty(row):
                skipped += 1
                continue

            try:
                action = resolve_action(row)
                if action is Action.IGNORE:
                    skipped += 1
                    continue
                processed += 1
                validation_errors = validate_ticket_row(
                    row, allowed_projects, legacy_default_project
                )
                if validation_errors:
                    raise TicketValidationError("; ".join(validation_errors))

                row["Project"] = validate_project_key(
                    row.get("Project"), allowed_projects, legacy_default_project
                )
                if not is_blank(row.get("IssueType")):
                    row["IssueType"] = normalize_issue_type(row.get("IssueType"))

                if action is Action.UPDATE:
                    jira_key = str(row.get("JiraKey")).strip().upper()
                    payload = build_issue_payload(
                        config,
                        row,
                        include_project=False,
                        include_issue_type=False,
                    )
                    result = process_update(config, jira_key, payload, dry_run)
                    if dry_run:
                        continue
                    if result == "NO_CHANGES":
                        _record_success(
                            sheet,
                            row_index,
                            headers,
                            str(config["JIRA_URL"]),
                            jira_key,
                            result,
                        )
                        continue
                    _record_success(
                        sheet,
                        row_index,
                        headers,
                        str(config["JIRA_URL"]),
                        jira_key,
                        result,
                    )
                    print(f"  [UPDATE] Zeile {row_index}: {jira_key}")
                    updated += 1
                    continue

                payload = build_issue_payload(config, row)
                if dry_run:
                    print(f"  [DRY RUN] Zeile {row_index}: CREATE validiert")
                    continue

                result = create_jira_issue(config, payload)
                jira_key = str(result.get("key") or "").strip().upper()
                if not jira_key:
                    raise RuntimeError(
                        "Jira-Erstellung lieferte keinen Jira Key; keine Wiederholung."
                    )
                _record_success(
                    sheet,
                    row_index,
                    headers,
                    str(config["JIRA_URL"]),
                    jira_key,
                    "CREATED",
                )
                print(f"  [OK] Zeile {row_index}: {jira_key}")
                created += 1
            except _EXPECTED_ROW_ERRORS as error:
                failed += 1
                message = str(error)
                _record_error(sheet, row_index, headers, message)
                print(f"  [FEHLER] Zeile {row_index}: {message}")

        if not dry_run:
            workbook.save(excel_file)
            print(f"\nExcel gespeichert: {excel_file}")
    finally:
        workbook.close()

    if not dry_run:
        reapply_dropdowns(excel_file)

    print()
    print("-" * 50)
    print(f"Verarbeitet:   {processed}")
    print(f"Erstellt:      {created}")
    print(f"Aktualisiert:  {updated}")
    print(f"Fehler:        {failed}")
    print(f"Uebersprungen: {skipped}")
    print(f"Dry Run:       {dry_run}")
    print("-" * 50)


if __name__ == "__main__":
    try:
        main()
    except (ConfigError, KeyError, OSError, RuntimeError, ValueError) as error:
        print(f"\nAbbruch: {error}", file=sys.stderr)
        sys.exit(1)
