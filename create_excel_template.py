"""
Erzeugt eine leere tickets.xlsx mit den korrekten Spaltenheadern.
Einmal ausführen, danach nicht mehr nötig.
"""

from openpyxl import Workbook

COLUMNS = [
    "Summary",
    "Description",
    "IssueType",
    "Priority",
    "Assignee",
    "Labels",
    "Components",
    "DueDate",
    "Status",
    "JiraKey",
    "ErrorMessage",
    "CreatedAt",
    "ExternalId",
]

EXAMPLE_ROW = [
    "Beispiel: Login-Bug auf iOS",
    "Nutzer kann sich nach Update nicht mehr einloggen.",
    "Bug",
    "High",
    "",  # Assignee - Jira-Username eintragen
    "frontend,urgent",
    "",  # Components - exakter Component-Name aus Jira (Pflichtfeld in DAH)
    "2026-06-15",
    "NEU",
    "",  # JiraKey - wird vom Skript befüllt
    "",  # ErrorMessage
    "",  # CreatedAt
    "EXCEL-001",
]


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"

    # Header
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Beispielzeile
    for col_idx, value in enumerate(EXAMPLE_ROW, start=1):
        ws.cell(row=2, column=col_idx, value=value)

    # Spaltenbreiten anpassen
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 10
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 40
    ws.column_dimensions["L"].width = 20
    ws.column_dimensions["M"].width = 14

    wb.save("tickets.xlsx")
    print("tickets.xlsx erstellt mit Beispielzeile.")


if __name__ == "__main__":
    main()
