from contextlib import redirect_stdout
from datetime import date, datetime
from io import StringIO
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest
from unittest.mock import Mock, patch

from openpyxl import Workbook, load_workbook

import create_tickets
from create_tickets import (
    build_issue_payload,
    create_jira_issue,
    format_due_date,
    process_update,
)


FAKE_TOKEN = "offline-only-secret-token"
BASE_CONFIG = {
    "JIRA_URL": "https://jira.example.test",
    "JIRA_TOKEN": FAKE_TOKEN,
    "PROJECT_KEY": "DAH",
    "ALLOWED_PROJECTS": ("DAH",),
    "EXCEL_FILE": "tickets.xlsx",
    "DRY_RUN": "true",
    "CURRENT_USER_NAME": "test-user",
}


def response(status, payload=None, text="", headers=None):
    result = Mock()
    result.status_code = status
    result.headers = headers or {}
    result.text = text
    result.json.return_value = payload
    return result


class CreateTicketsTests(unittest.TestCase):
    def test_create_payload_preserves_reporter_automation(self):
        row = {
            "Action": "CREATE",
            "Project": "DAH",
            "IssueType": "Bug",
            "Summary": "Summary",
            "Description": "Text",
            "Components": "Backend, Frontend",
            "Labels": "one, two",
            "Reporter": "must-not-be-sent",
            "DueDate": date(2026, 7, 21),
        }
        fields = build_issue_payload(BASE_CONFIG, row)["fields"]
        self.assertEqual(fields["project"], {"key": "DAH"})
        self.assertEqual(fields["issuetype"], {"name": "Bug"})
        self.assertNotIn("reporter", fields)
        self.assertEqual(fields["duedate"], "2026-07-21")
        self.assertEqual(
            fields["components"], [{"name": "Backend"}, {"name": "Frontend"}]
        )

    def test_update_payload_omits_locked_fields_and_supports_clear(self):
        row = {
            "Project": "OTHER",
            "IssueType": "Epic",
            "Summary": "",
            "Description": "<CLEAR>",
            "Priority": "",
            "Assignee": "Unassigned",
            "Labels": "<CLEAR>",
            "Components": "<CLEAR>",
            "DueDate": "<CLEAR>",
            "Reporter": "another-user",
            "JiraStatus": "Done",
        }
        fields = build_issue_payload(
            BASE_CONFIG, row, include_project=False, include_issue_type=False
        )["fields"]
        self.assertNotIn("project", fields)
        self.assertNotIn("issuetype", fields)
        self.assertNotIn("reporter", fields)
        self.assertNotIn("status", fields)
        self.assertNotIn("summary", fields)
        self.assertIsNone(fields["description"])
        self.assertIsNone(fields["assignee"])
        self.assertEqual(fields["labels"], [])
        self.assertEqual(fields["components"], [])
        self.assertIsNone(fields["duedate"])

    def test_dates_are_serialized_for_jira(self):
        self.assertEqual(format_due_date(datetime(2026, 7, 21, 12, 30)), "2026-07-21")
        self.assertEqual(format_due_date("21.07.26"), "2026-07-21")
        with self.assertRaisesRegex(ValueError, "Due Date"):
            format_due_date("tomorrow")

    @patch("create_tickets.requests.post")
    def test_create_uses_exactly_one_post(self, post):
        post.return_value = response(201, {"key": "DAH-1"})
        result = create_jira_issue(BASE_CONFIG, {"fields": {"summary": "x"}})
        self.assertEqual(result["key"], "DAH-1")
        post.assert_called_once()

    @patch("create_tickets.requests.post")
    def test_create_rejects_unexpected_response_shape(self, post):
        post.return_value = response(201, ["unexpected"])
        with self.assertRaisesRegex(RuntimeError, "unerwartetes Antwortformat"):
            create_jira_issue(BASE_CONFIG, {"fields": {"summary": "x"}})

    @patch("create_tickets.requests.post")
    def test_http_failure_does_not_expose_response_or_token(self, post):
        post.return_value = response(
            500,
            {"token": FAKE_TOKEN},
            text=f"full sensitive response {FAKE_TOKEN}",
        )
        with self.assertRaises(RuntimeError) as context:
            create_jira_issue(BASE_CONFIG, {"fields": {"summary": "x"}})
        message = str(context.exception)
        self.assertIn("HTTP 500", message)
        self.assertNotIn(FAKE_TOKEN, message)
        self.assertNotIn("full sensitive response", message)

    @patch("create_tickets.requests.put")
    @patch("create_tickets.requests.get")
    def test_update_dry_run_reads_once_and_never_writes(self, get, put):
        get.return_value = response(
            200, {"fields": {"summary": "old", "updated": "2026-07-21T10:00:00"}}
        )
        output = StringIO()
        with redirect_stdout(output):
            result = process_update(
                BASE_CONFIG,
                "DAH-1",
                {"fields": {"summary": "new"}},
                dry_run=True,
            )
        self.assertEqual(result, "DRY_RUN")
        self.assertEqual(get.call_count, 1)
        put.assert_not_called()
        self.assertNotIn(FAKE_TOKEN, output.getvalue())

    @patch("create_tickets.requests.put")
    @patch("create_tickets.requests.get")
    def test_confirmed_update_rechecks_timestamp_then_writes_changed_fields(
        self, get, put
    ):
        get.side_effect = [
            response(
                200,
                {
                    "fields": {
                        "summary": "old",
                        "description": "same",
                        "updated": "2026-07-21T10:00:00",
                    }
                },
            ),
            response(200, {"fields": {"updated": "2026-07-21T10:00:00"}}),
        ]
        put.return_value = response(204)
        result = process_update(
            BASE_CONFIG,
            "DAH-1",
            {"fields": {"summary": "new", "description": "same"}},
            dry_run=False,
            confirmation=lambda _key, _fields: True,
        )
        self.assertEqual(result, "UPDATED")
        self.assertEqual(get.call_count, 2)
        put.assert_called_once()
        self.assertEqual(
            put.call_args.kwargs["json"], {"fields": {"summary": "new"}}
        )

    @patch("create_tickets.requests.put")
    @patch("create_tickets.requests.get")
    def test_concurrent_change_blocks_update(self, get, put):
        get.side_effect = [
            response(200, {"fields": {"summary": "old", "updated": "t1"}}),
            response(200, {"fields": {"updated": "t2"}}),
        ]
        with self.assertRaisesRegex(RuntimeError, "zwischenzeitlich"):
            process_update(
                BASE_CONFIG,
                "DAH-1",
                {"fields": {"summary": "new"}},
                dry_run=False,
                confirmation=lambda _key, _fields: True,
            )
        put.assert_not_called()

    @patch("create_tickets.requests.put")
    @patch("create_tickets.requests.get")
    def test_unconfirmed_update_never_writes(self, get, put):
        get.return_value = response(
            200, {"fields": {"summary": "old", "updated": "t1"}}
        )
        with self.assertRaisesRegex(RuntimeError, "nicht bestaetigt"):
            process_update(
                BASE_CONFIG,
                "DAH-1",
                {"fields": {"summary": "new"}},
                dry_run=False,
                confirmation=lambda _key, _fields: False,
            )
        put.assert_not_called()

    def test_main_isolates_invalid_row_and_hyperlinks_success(self):
        with TemporaryDirectory() as directory:
            path = Path(directory) / "tickets.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Tickets"
            sheet.append(
                [
                    "Action",
                    "Project",
                    "IssueType",
                    "Summary",
                    "Components",
                    "JiraKey",
                    "Result",
                    "ErrorMessage",
                ]
            )
            sheet.append(["CREATE", "DAH", "Story", "Invalid", "", "", "", ""])
            sheet.append(["CREATE", "DAH", "Incident", "Valid", "", "", "", ""])
            workbook.save(path)
            workbook.close()

            config = dict(BASE_CONFIG, EXCEL_FILE=str(path), DRY_RUN="false")
            output = StringIO()
            with (
                patch("create_tickets.load_config", return_value=config),
                patch("create_tickets.test_jira_connection"),
                patch("create_tickets.test_project_access"),
                patch(
                    "create_tickets.create_jira_issue",
                    return_value={"key": "DAH-99"},
                ) as create,
                redirect_stdout(output),
            ):
                create_tickets.main()

            result = load_workbook(path)
            result_sheet = result["Tickets"]
            self.assertEqual(result_sheet["G2"].value, "ERROR")
            self.assertIn("Component/s fehlt", result_sheet["H2"].value)
            self.assertEqual(result_sheet["F3"].value, "DAH-99")
            self.assertEqual(result_sheet["G3"].value, "CREATED")
            self.assertEqual(result_sheet["A3"].value, "IGNORE")
            self.assertEqual(
                result_sheet["F3"].hyperlink.target,
                "https://jira.example.test/browse/DAH-99",
            )
            result.close()
            create.assert_called_once()
            self.assertNotIn(FAKE_TOKEN, output.getvalue())


if __name__ == "__main__":
    unittest.main()
